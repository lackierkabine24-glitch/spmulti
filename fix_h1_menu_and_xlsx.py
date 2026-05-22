#!/usr/bin/env python3
import csv
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
PUBLIC = ROOT / 'public'
SRC_PAGES = ROOT / 'src' / 'pages'
MENU_XLSX = PUBLIC / 'menusp.xlsx'
MENU_CSV = PUBLIC / 'menusp.csv'
H1_XLSX = PUBLIC / 'h1-headings.xlsx'

if __name__ == '__main__':
    # 1. Clean public/menusp.csv from the XLSX source
    if MENU_XLSX.exists():
        wb = openpyxl.load_workbook(MENU_XLSX)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            first = ''
            second = ''
            if row[0] is not None:
                first = str(row[0]).strip()
            if len(row) > 1 and row[1] is not None:
                second = str(row[1]).strip()
            if not first and not second:
                continue
            if first.lower() in ('table 1', 'menu link') and not second:
                continue
            rows.append([first, second])
        with open(MENU_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        print(f'✅ Cleaned menu CSV written to {MENU_CSV}')
    else:
        print(f'⚠️  Missing menu XLSX source: {MENU_XLSX}')

    # 2. Update pages with H1 if missing
    page_rows = []
    pattern_title = re.compile(r'const\s+title\s*=\s*["\'](.+?)["\'];')
    pattern_layout = re.compile(r'(<Layout[^>]*title=\{title\}[^>]*>)(\s*)', re.MULTILINE)
    files = sorted(SRC_PAGES.rglob('index.astro'))
    inserted_files = []
    for page in files:
        text = page.read_text(encoding='utf-8')
        has_h1 = '<h1' in text.lower()
        title_match = pattern_title.search(text)
        title = title_match.group(1).strip() if title_match else ''
        if not title:
            rel = page.relative_to(SRC_PAGES)
            if str(rel) == 'index.astro':
                title = 'Salespeare'
            else:
                parent = rel.parent.name
                title = parent.replace('-', ' ').replace('_', ' ').title()
        if not has_h1:
            m = pattern_layout.search(text)
            if m:
                insert = f"{m.group(1)}{m.group(2)}  <h1>{title}</h1>\n"
                text = text[:m.start()] + insert + text[m.end():]
                page.write_text(text, encoding='utf-8')
                inserted_files.append(str(page.relative_to(ROOT)))
        rel = page.relative_to(SRC_PAGES)
        if str(rel) == 'index.astro':
            url = '/'
        else:
            url = '/' + str(rel.parent).replace('\\', '/')
        page_rows.append((str(rel), url, title))

    print(f'✅ Inserted H1 into {len(inserted_files)} page(s)')
    for path in inserted_files[:20]:
        print(f'  - {path}')
    if len(inserted_files) > 20:
        print(f'  ...and {len(inserted_files)-20} more')

    # 3. Create H1 spreadsheet at /public/h1-headings.xlsx
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = 'H1 Headings'
    ws2.append(['page', 'url', 'h1'])
    for row in page_rows:
        ws2.append(row)
    wb2.save(H1_XLSX)
    print(f'✅ H1 spreadsheet created at {H1_XLSX}')
